#!/usr/bin/env python3
"""
apply_minmax_recommendations.py

Applies the Min/Max Recommendation Report for a single store to the vendor
Rules Matrix, writing the recommended values into that store's `_Min` and
`_Max` columns (matched by SKU).

The recommendation workbook stores its recommendations as live formulas
(Recommended Min / Recommended Max) that reference a small set of config
cells (report window, vendor lead time, safety stock buffer, order cycle).
Those formulas are recalculated here in Python rather than relying on
cached values, since a file saved by openpyxl/scripted tools often has no
cached formula results.

Usage:
    python apply_minmax_recommendations.py

Just run it -- it will ask you for the Rules Matrix path, the
recommendations file path, the store code, and a few yes/no options.

Store codes must match the column prefixes already used in the Rules
Matrix header row, e.g. CM, CVM, CC, DTD, MF, LB, LF, PP, SP, SH, SS, HQ.
"""

import math
import os

import openpyxl


def find_recommendation_config(ws):
    """Scan column A for the labeled config cells and return their values."""
    labels = {
        "Report window (days):": "window",
        "Vendor lead time (days):": "lead_time",
        "Safety stock buffer (days):": "buffer",
        "Order cycle (days):": "order_cycle",
    }
    config = {}
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=2):
        label_cell, value_cell = row[0], row[1]
        label = label_cell.value
        if isinstance(label, str) and label.strip() in labels:
            config[labels[label.strip()]] = value_cell.value

    missing = [key for key in (
        "window", "lead_time", "buffer", "order_cycle") if key not in config]
    if missing:
        raise ValueError(
            f"Could not find config values for: {missing}. "
            "Check that the recommendations sheet layout hasn't changed."
        )
    return config


def find_recommendation_table(ws):
    """Locate the header row (Item Name / SKU / ... ) and return its index."""
    for row in ws.iter_rows(min_row=1, max_row=20):
        values = [cell.value for cell in row]
        if len(values) >= 2 and values[0] == "Item Name" and values[1] == "SKU":
            return row[0].row
    raise ValueError(
        "Could not find the 'Item Name' / 'SKU' header row in the recommendations sheet.")


def roundup(value, digits=0):
    """Excel ROUNDUP behavior for non-negative numbers."""
    factor = 10 ** digits
    return math.ceil(value * factor) / factor


def compute_recommendations(rec_path):
    """
    Read the recommendations workbook and return a dict:
        { sku_str: {"min": int, "max": int, "item_name": str} }
    Recreates the sheet's own formulas rather than trusting cached values.
    """
    wb = openpyxl.load_workbook(rec_path, data_only=True)
    ws = wb.worksheets[0]

    config = find_recommendation_config(ws)
    window = config["window"]
    lead_time = config["lead_time"]
    buffer_days = config["buffer"]
    order_cycle = config["order_cycle"]

    header_row = find_recommendation_table(ws)

    recs = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_col=5):
        item_name, sku, gtin, on_hand, sold = (c.value for c in row)
        if item_name is None and sku is None:
            continue
        if sku is None:
            continue
        sku = str(sku).strip()
        sold = sold or 0

        if sold == 0:
            rec_min, rec_max = 0, 0
        else:
            daily_rate = round(sold / window, 2)
            rec_min = int(roundup(daily_rate * (lead_time + buffer_days)))
            rec_max = int(roundup(rec_min + daily_rate * order_cycle))

        recs[sku] = {"min": rec_min, "max": rec_max, "item_name": item_name}

    return recs


def apply_to_rules_matrix(matrix_path, recs, store, include_dno=False):
    """
    Load the Rules Matrix, update {store}_Min / {store}_Max for matched SKUs,
    and return (workbook, summary_dict). Does not save.
    """
    wb = openpyxl.load_workbook(matrix_path)
    ws = wb.worksheets[0]

    header = [cell.value for cell in ws[1]]
    col_index = {name: idx + 1 for idx,
                 name in enumerate(header)}  # 1-based for openpyxl

    dno_col = f"{store}_DNO"
    min_col = f"{store}_Min"
    max_col = f"{store}_Max"

    missing_cols = [c for c in (
        dno_col, min_col, max_col) if c not in col_index]
    if missing_cols:
        valid_stores = sorted(
            {h.rsplit("_", 1)[0] for h in header if h and h.endswith("_Min")})
        raise ValueError(
            f"Store '{store}' not found (missing columns: {missing_cols}). "
            f"Valid store codes in this Rules Matrix: {valid_stores}"
        )

    sku_col = col_index["SKU"]

    updated, skipped_dno, not_in_matrix = [], [], []
    matched_skus = set()

    for row_cells in ws.iter_rows(min_row=2):
        sku_cell = row_cells[sku_col - 1]
        sku = sku_cell.value
        if sku is None:
            continue
        sku = str(sku).strip()
        if sku not in recs:
            continue

        matched_skus.add(sku)
        rec = recs[sku]
        row_num = sku_cell.row

        dno_value = ws.cell(row=row_num, column=col_index[dno_col]).value
        is_dno = bool(dno_value) and str(
            dno_value).strip().upper() not in ("FALSE", "0", "")

        if is_dno and not include_dno:
            skipped_dno.append({"sku": sku, "item_name": rec["item_name"]})
            continue

        old_min = ws.cell(row=row_num, column=col_index[min_col]).value
        old_max = ws.cell(row=row_num, column=col_index[max_col]).value

        ws.cell(row=row_num, column=col_index[min_col]).value = rec["min"]
        ws.cell(row=row_num, column=col_index[max_col]).value = rec["max"]

        updated.append({
            "sku": sku,
            "item_name": rec["item_name"],
            "old_min": old_min, "new_min": rec["min"],
            "old_max": old_max, "new_max": rec["max"],
        })

    not_in_matrix = [{"sku": sku, "item_name": r["item_name"]}
                     for sku, r in recs.items() if sku not in matched_skus]

    summary = {
        "updated": updated,
        "skipped_dno": skipped_dno,
        "not_in_matrix": not_in_matrix,
    }
    return wb, summary


def print_summary(store, summary, dry_run):
    updated = summary["updated"]
    skipped_dno = summary["skipped_dno"]
    not_in_matrix = summary["not_in_matrix"]

    verb = "Would update" if dry_run else "Updated"
    print(f"\n=== {store} Min/Max {'preview' if dry_run else 'apply'} ===")
    print(f"{verb} {len(updated)} item(s):")
    for u in updated:
        changed = (u["old_min"] != u["new_min"]) or (
            u["old_max"] != u["new_max"])
        marker = "*" if changed else " "
        print(f"  {marker} [{u['sku']}] {u['item_name']}: "
              f"Min {u['old_min']} -> {u['new_min']}, Max {u['old_max']} -> {u['new_max']}")

    if skipped_dno:
        print(
            f"\nSkipped {len(skipped_dno)} item(s) marked DNO for {store} (use --include-dno to override):")
        for s in skipped_dno:
            print(f"    [{s['sku']}] {s['item_name']}")

    if not_in_matrix:
        print(
            f"\n{len(not_in_matrix)} SKU(s) in the recommendations file were not found in the Rules Matrix:")
        for n in not_in_matrix:
            print(f"    [{n['sku']}] {n['item_name']}")


def prompt_path(message):
    """Ask for a file path and keep asking until it exists."""
    while True:
        path = input(message).strip().strip('"')
        if os.path.isfile(path):
            return path
        print(f"  Can't find that file: {path}\n")


def prompt_yes_no(message, default=False):
    suffix = " [Y/n]: " if default else " [y/N]: "
    answer = input(message + suffix).strip().lower()
    if not answer:
        return default
    return answer.startswith("y")


def main():
    print("=== Apply Min/Max Recommendations ===\n")

    rules_matrix_path = prompt_path("Path to the Rules Matrix .xlsx: ")
    recommendations_path = prompt_path(
        "Path to the store's Min/Max Recommendation Report .xlsx: ")
    store = input(
        "Store code (matches the column prefix in the Rules Matrix, e.g. CM): ").strip().upper()

    include_dno = prompt_yes_no(
        "Also update items flagged DNO for this store?", default=False)
    dry_run = prompt_yes_no(
        "Dry run only (preview, don't save)?", default=False)

    out_path = None
    if not dry_run:
        in_place = prompt_yes_no(
            "Overwrite the original Rules Matrix file in place?", default=False)
        if in_place:
            out_path = rules_matrix_path
        else:
            default_out = f"{rules_matrix_path.rsplit('.', 1)[0]}_updated.xlsx"
            entered = input(
                f"Output file path [{default_out}]: ").strip().strip('"')
            out_path = entered or default_out

    recs = compute_recommendations(recommendations_path)
    wb, summary = apply_to_rules_matrix(
        rules_matrix_path, recs, store, include_dno=include_dno)

    print_summary(store, summary, dry_run)

    if dry_run:
        print("\nDry run only -- no file was saved.")
        return

    wb.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
