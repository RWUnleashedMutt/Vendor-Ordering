#!/usr/bin/env python3
"""
Min/Max Recommendation Report Generator
========================================

Reads a Square "Inventory sell-through" export, cleans it down to one row
per item/variation, and produces a Min/Max stocking-level recommendation
workbook.

Works with the export from any single store - it just asks which store
the file came from and stamps that on the report, so you can run this
against each location's export and get a comparable set of reports.

Usage:
    python minmax_recommendation.py

You'll be prompted for:
    - Store name
    - Path to the Square export (.xlsx)
    - Vendor lead time (days)
    - Safety stock buffer (days)
    - Order cycle - how often you place orders (days)

Output:
    A new .xlsx workbook named "<Store Name> - MinMax Recommendations
    <date>.xlsx" written next to the input file, with live formulas so
    you can tweak the assumption cells (lead time / safety / cycle) and
    everything recalculates in Excel.
"""

import sys
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Square's sell-through export covers a trailing 30-day window
REPORT_WINDOW_DAYS_DEFAULT = 30

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", bold=True)
LABEL_FONT = Font(name=FONT_NAME, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
NORMAL_FONT = Font(name=FONT_NAME)
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Input collection
# ---------------------------------------------------------------------------

def prompt(text, default=None, cast=str):
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"{text}{suffix}: ").strip()
        if not raw and default is not None:
            return default
        if not raw:
            print("This field is required.")
            continue
        try:
            return cast(raw)
        except ValueError:
            print(f"Please enter a valid {cast.__name__}.")


def clean_path(raw_path):
    """
    Strip whitespace and any surrounding quote characters. Copying a path
    from Windows Explorer ("Copy as path") or dragging a file into some
    terminals wraps it in double quotes - strip those (and single quotes)
    so the user doesn't have to edit them out by hand.
    """
    return raw_path.strip().strip('"').strip("'").strip()


def get_inputs():
    print("=" * 60)
    print("MIN / MAX RECOMMENDATION REPORT")
    print("=" * 60)
    store_name = prompt("Store name")

    file_path = clean_path(
        prompt("Path to the Square sell-through export (.xlsx)"))
    while not Path(file_path).expanduser().exists():
        print(f"File not found: {file_path}")
        file_path = clean_path(
            prompt("Path to the Square sell-through export (.xlsx)"))

    print("\nReorder assumptions (press Enter to accept the default):")
    lead_time = prompt("  Vendor lead time in days", default=7, cast=int)
    safety_days = prompt("  Safety stock buffer in days", default=7, cast=int)
    order_cycle = prompt(
        "  Order cycle - how often you reorder, in days", default=14, cast=int)

    return store_name, str(Path(file_path).expanduser()), lead_time, safety_days, order_cycle


# ---------------------------------------------------------------------------
# Cleaning
# ---------------------------------------------------------------------------

def clean_data(df):
    """
    Square exports two rows per item: an 'All' roll-up row (no GTIN) and one
    or more variation detail rows. Collapse each item down to just the
    row(s) that carry a GTIN, dropping the duplicate roll-up row. If an item
    has no GTIN on any row, fall back to keeping the row with a SKU instead
    of losing the item entirely.
    """
    df = df.copy()
    original_items = df["Item Name"].nunique()

    kept_groups = []
    no_gtin_items = []

    for name, g in df.groupby("Item Name", sort=False):
        if len(g) == 1:
            kept_groups.append(g)
            continue

        has_gtin = g[g["GTIN"].notna()]
        if len(has_gtin) > 0:
            kept_groups.append(has_gtin)
            continue

        has_sku = g[g["SKU"].notna()]
        if len(has_sku) > 0:
            kept_groups.append(has_sku.iloc[[0]])
        else:
            kept_groups.append(g.iloc[[0]])
        no_gtin_items.append(name)

    clean = pd.concat(kept_groups, ignore_index=True)

    info = {
        "original_items": original_items,
        "original_rows": len(df),
        "clean_rows": len(clean),
        "no_gtin_items": sorted(set(no_gtin_items)),
    }
    return clean, info


# ---------------------------------------------------------------------------
# Workbook build (values written; recommendation math done as live formulas)
# ---------------------------------------------------------------------------

def build_workbook(clean_df, store_name, source_file, lead_time, safety_days,
                   order_cycle, report_window, out_path):

    wb = Workbook()
    ws = wb.active
    ws.title = "Min-Max Recommendations"

    # --- Title block -------------------------------------------------
    ws["A1"] = f"Min/Max Recommendation Report - {store_name}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source file: {Path(source_file).name}"
    ws["A2"].font = NORMAL_FONT
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = NORMAL_FONT

    # --- Assumptions (editable, blue) --------------------------------
    ws["A5"] = "Report window (days):"
    ws["A5"].font = LABEL_FONT
    ws["B5"] = report_window
    ws["B5"].font = INPUT_FONT

    ws["A6"] = "Vendor lead time (days):"
    ws["A6"].font = LABEL_FONT
    ws["B6"] = lead_time
    ws["B6"].font = INPUT_FONT

    ws["A7"] = "Safety stock buffer (days):"
    ws["A7"].font = LABEL_FONT
    ws["B7"] = safety_days
    ws["B7"].font = INPUT_FONT

    ws["A8"] = "Order cycle (days):"
    ws["A8"].font = LABEL_FONT
    ws["B8"] = order_cycle
    ws["B8"].font = INPUT_FONT

    ws["D5"] = "Edit the blue cells above and the recommendations below recalculate."
    ws["D5"].font = Font(name=FONT_NAME, italic=True, size=9, color="808080")

    # --- Table header --------------------------------------------------
    header_row = 10
    headers = [
        "Item Name", "SKU", "GTIN", "On Hand", "Sold (period)",
        "Sell-through", "Daily Rate", "Recommended Min", "Recommended Max",
        "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center",
                                vertical="center", wrap_text=True)
        c.border = BORDER

    # --- Data rows -------------------------------------------------
    first_data_row = header_row + 1
    clean_df = clean_df.sort_values("Item Name").reset_index(drop=True)

    for i, row in clean_df.iterrows():
        r = first_data_row + i
        on_hand = int(row["On hand"]) if pd.notna(row["On hand"]) else 0
        sold = int(row["Sold"]) if pd.notna(row["Sold"]) else 0
        sell_through = row["Sell-through"] if pd.notna(
            row["Sell-through"]) else None
        gtin = row["GTIN"]
        gtin_str = "" if pd.isna(gtin) else str(int(gtin))
        sku = "" if pd.isna(row["SKU"]) else str(row["SKU"])

        ws.cell(row=r, column=1, value=row["Item Name"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=sku).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=gtin_str).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=on_hand).font = NORMAL_FONT
        ws.cell(row=r, column=5, value=sold).font = NORMAL_FONT

        st_cell = ws.cell(row=r, column=6, value=sell_through)
        st_cell.font = NORMAL_FONT
        st_cell.number_format = "0.0%"

        # Daily Rate = Sold / report window (assumption cell $B$5)
        ws.cell(row=r, column=7,
                value=f"=ROUND(E{r}/$B$5,2)").font = NORMAL_FONT

        # Recommended Min = ceil(DailyRate * (lead time + safety days)); 0 if no sales
        ws.cell(row=r, column=8,
                value=f"=IF(E{r}=0,0,ROUNDUP(G{r}*($B$6+$B$7),0))").font = NORMAL_FONT

        # Recommended Max = Min + DailyRate * order cycle; 0 if no sales
        ws.cell(row=r, column=9,
                value=f"=IF(E{r}=0,0,ROUNDUP(H{r}+G{r}*$B$8,0))").font = NORMAL_FONT

        # Notes - flags built from plain IF/AND logic (LibreOffice-safe)
        notes_formula = (
            f'=TRIM('
            f'IF(AND(E{r}=0,D{r}>0),"Dead stock - no sales in period. ","")&'
            f'IF(AND(D{r}<=0,E{r}>0),"STOCKOUT - reorder now. ","")&'
            f'IF(AND(I{r}>0,D{r}>I{r}),"Overstocked vs Max. ","")&'
            f'IF(AND(E{r}>0,D{r}<H{r}),"Below Min - reorder. ","")'
            f')'
        )
        notes_cell = ws.cell(row=r, column=10, value=notes_formula)
        notes_cell.font = NORMAL_FONT
        notes_cell.alignment = Alignment(wrap_text=True)

        for col in range(1, 11):
            ws.cell(row=r, column=col).border = BORDER

    last_row = first_data_row + len(clean_df) - 1

    # --- Formatting ------------------------------------------------
    widths = [34, 12, 14, 10, 12, 12, 11, 16, 16, 42]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:J{last_row}"

    # Conditional-format the Notes column so flagged rows stand out
    from openpyxl.formatting.rule import FormulaRule
    ws.conditional_formatting.add(
        f"A{first_data_row}:J{last_row}",
        FormulaRule(formula=[f'$J{first_data_row}<>""'], fill=FLAG_FILL),
    )

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    store_name, file_path, lead_time, safety_days, order_cycle = get_inputs()

    print("\nLoading file...")
    # SKU is read as a string dtype so numeric-looking SKUs don't get
    # silently coerced to float64 by pandas (which would turn e.g. "12345"
    # into 12345.0, and later str(12345.0) -> "12345.0" downstream).
    df = pd.read_excel(file_path, dtype={"SKU": str})

    required_cols = {"Item Name", "Variation Name", "GTIN", "SKU", "Sell-through",
                     "On hand", "Sold"}
    missing = required_cols - set(df.columns)
    if missing:
        print(f"ERROR: file is missing expected columns: {missing}")
        sys.exit(1)

    print("Cleaning data (one row per item, dropping no-GTIN duplicate rows)...")
    clean_df, info = clean_data(df)
    print(
        f"  Original: {info['original_items']} items across {info['original_rows']} rows")
    print(f"  Cleaned to: {len(clean_df)} rows")
    if info["no_gtin_items"]:
        print(f"  Note: {len(info['no_gtin_items'])} items have no GTIN on any row; "
              f"kept using SKU/name instead so nothing was dropped.")

    out_name = f"./Data/SLT Reports/{store_name} - MinMax Recommendations {datetime.now().strftime('%Y-%m-%d')}.xlsx"
    # Write next to where the script is run from (avoids failing if the
    # input file lives in a read-only or synced folder).
    out_path = str(Path.cwd() / out_name)

    print("Building recommendation workbook...")
    build_workbook(clean_df, store_name, file_path, lead_time, safety_days,
                   order_cycle, REPORT_WINDOW_DAYS_DEFAULT, out_path)

    print(f"\nDone -> {out_path}")


if __name__ == "__main__":
    main()
